import pandas as pd
from datetime import date
from openpyxl import load_workbook

ARCHIVO = "BioMonitor.xlsx"

# ─────────────────────────────────────────────
#  CARGA DE DATOS DESDE EXCEL
# ─────────────────────────────────────────────
xl            = pd.read_excel(ARCHIVO, sheet_name=None, header=None)
df_usuarios   = pd.read_excel(ARCHIVO, sheet_name="Usuario")
df_malestares = xl["Malestares"]

# ── Parsear hoja Plantas (estructura en columnas por tipo) ──
_plantas_raw = xl["Plantas"]
TIPOS = ["Raíces", "Bulbos", "Tallo", "Flores", "Fruto", "Semillas"]
COL   = [1, 2, 3, 4, 5, 6]          # columnas 1-6 (0-indexed) de la hoja

# Índices de fila con datos clave (0-indexed, header=None)
# Fila 0: "Tipo"   Fila 1+: especies   Fila -5: sol  -4: agua_veces  -3: agua_min  -2: plantas_m2  -1: separación
_fila_sol       = _plantas_raw[_plantas_raw.iloc[:, 0] == "Requisitos sol"].index[0]
_fila_aw        = _plantas_raw[_plantas_raw.iloc[:, 0].str.startswith("Requisitos agua (veces", na=False)].index[0]
_fila_am        = _plantas_raw[_plantas_raw.iloc[:, 0].str.startswith("Requisitos agua (minutos", na=False)].index[0]
_fila_pm2       = _plantas_raw[_plantas_raw.iloc[:, 0].str.startswith("Espacio (plantas", na=False)].index[0]
_fila_sep       = _plantas_raw[_plantas_raw.iloc[:, 0].str.startswith("Espacio (separación", na=False)].index[0]

# Construir diccionario tipo → info
INFO_TIPO = {}
for tipo, col in zip(TIPOS, COL):
    especies = []
    for i in range(1, _fila_sol):
        v = _plantas_raw.iat[i, col]
        if pd.notna(v):
            especies.append(str(v).strip().lower())
    INFO_TIPO[tipo] = {
        "especies":    especies,
        "sol":         str(_plantas_raw.iat[_fila_sol,  col]),
        "agua_veces":  str(_plantas_raw.iat[_fila_aw,   col]),
        "agua_min":    str(_plantas_raw.iat[_fila_am,   col]),
        "plantas_m2":  str(_plantas_raw.iat[_fila_pm2,  col]),
        "separacion":  str(_plantas_raw.iat[_fila_sep,  col]),
    }

# Mapa inverso planta → tipo
PLANTA_TIPO = {}
for tipo, info in INFO_TIPO.items():
    for esp in info["especies"]:
        PLANTA_TIPO[esp] = tipo


# ─────────────────────────────────────────────
#  MÓDULO DE AUTENTICACIÓN
# ─────────────────────────────────────────────
def login():
    print("╔══════════════════════════════╗")
    print("║   Bienvenido a BioMonitor    ║")
    print("╚══════════════════════════════╝\n")
    intentos = 3
    while intentos > 0:
        usuario    = input("Usuario: ").strip()
        contrasena = input("Contraseña: ").strip()
        match = df_usuarios[
            (df_usuarios["usuario"] == usuario) &
            (df_usuarios["contraseña"].astype(str) == contrasena)
        ]
        if not match.empty:
            print(f"\n✅ Bienvenido, {usuario}!\n")
            return usuario
        else:
            intentos -= 1
            print(f"❌ Datos incorrectos. Intentos restantes: {intentos}\n")
    print("Demasiados intentos fallidos. Saliendo…")
    exit()


# ─────────────────────────────────────────────
#  MÓDULO DE DIAGNÓSTICO DE MALESTARES
# ─────────────────────────────────────────────
def malestar():
    preguntas = [
        "¿Tu planta está mordida?",
        "¿Las mordidas siguen un patrón blanco?",
        "¿Las hojas de tu planta perdieron color?",
        "¿Tus hojas tienen daños de succión?",
        "¿Tus hojas tienen puntos amarillos o blancos?",
        "¿Has visto pequeños arácnidos rojos?",
        "¿Has visto insectos cafés o negros tipo arroz?",
        "¿Tus hojas tienen una capa blanca tipo telaraña?",
        "¿Tus hojas se ven esqueletizadas?",
        "¿Has visto bolitas negras (excremento)?",
        "¿Hay bichos verdes debajo de las hojas?",
        "¿Tu planta tiene necrosis (partes negras)?",
        "¿Tienes hojas amarillentas?",
        "¿Las hojas están tristes y crujientes?",
        "¿La lesión tiene bordes claros?",
        "¿Las lesiones parecen mosaico?"
    ]
    casos = [
        (["Si","Si","No","No","No","No","No","No","No","No","No","No","No","No","No","No"],
         "Escarabajos, moscas y polillas", "Agua de vidrio"),
        (["No","No","Si","Si","Si","Si","No","Si","No","No","No","No","No","No","No","No"],
         "Ácaros", "Concentrado de chile"),
        (["No","No","Si","Si","Si","No","Si","No","No","No","No","No","No","No","No","No"],
         "Thrips", "Agua de vidrio"),
        (["Si","No","No","No","No","No","No","No","Si","Si","No","No","No","No","No","No"],
         "Orugas, escarabajos o moscas", "Agua con vinagre"),
        (["No","No","No","No","No","No","No","No","No","No","Si","No","No","No","No","No"],
         "Áfidos", "Podar, agua con vinagre y riego a presión"),
        (["No","No","Si","No","No","No","No","No","No","No","No","Si","Si","No","No","Si"],
         "Virus", "Talar"),
        (["No","No","No","No","No","No","No","No","No","No","No","Si","No","Si","No","No"],
         "Bacteria", "Talar"),
        (["No","No","No","No","No","No","No","No","No","No","No","Si","No","No","Si","No"],
         "Hongo", "Podar y aplicar fungicida de ajo y vinagre")
    ]

    respuestas_usuario = []
    print("Responde con 'Si' o 'No'\n")
    for pregunta in preguntas:
        while True:
            r = input(pregunta + " ").strip().capitalize()
            if r in ["Si", "No"]:
                respuestas_usuario.append(r)
                break
            else:
                print("Por favor responde solo 'Si' o 'No'.")

    encontrado = False
    diagnostico_resultado = ""
    for condiciones, diagnostico, solucion in casos:
        if respuestas_usuario == condiciones:
            print("\nDiagnóstico:", diagnostico)
            print("Solución:", solucion)
            diagnostico_resultado = f"{diagnostico} → {solucion}"
            encontrado = True
            break

    if not encontrado:
        print("\nNo se encontró un diagnóstico exacto.")
        print("Revisa las respuestas o consulta a un especialista.")
        diagnostico_resultado = "Sin diagnóstico exacto"

    return diagnostico_resultado


# ─────────────────────────────────────────────
#  MÓDULO DE CÁLCULO POR TIPO DE PLANTA
# ─────────────────────────────────────────────
def calcular_plantas():
    h = float(input("¿Qué tan ancho es tu huerto (en m)? "))
    l = float(input("¿Qué tan largo es tu huerto (en m)? "))
    area = round(h * l, 2)
    print(f"El área de tu huerto es {area} m²\n")

    opciones = list(INFO_TIPO.keys())
    print("---Menú de selección del tipo de tu planta---")
    for i, t in enumerate(opciones, 1):
        print(f"{i}. {t}")
    print("0. Salir")

    opcion = int(input("\nElige! "))
    if opcion == 0 or opcion > len(opciones):
        print("Opción inválida.")
        return

    tipo = opciones[opcion - 1]
    info = INFO_TIPO[tipo]

    rango   = info["plantas_m2"].replace("–", "-").split("-")
    sep_cm  = int("".join(filter(str.isdigit, info["separacion"])))
    min_p   = int(float(rango[0].strip()) * area)
    max_p   = int(float(rango[1].strip()) * area)
    surcos  = max(1, int((h * 100) / sep_cm))

    print(f"\nEn tu huerto caben entre {min_p} y {max_p} plantas de tipo {tipo}.")
    print(f"Caben aproximadamente {surcos} surcos.")
    print(f"Entre {min_p // surcos} y {max_p // surcos} plantas por surco.")


# ─────────────────────────────────────────────
#  MÓDULO DE RECETARIO
# ─────────────────────────────────────────────
def recetario():
    print("\n╔══════════════════════════════╗")
    print("║      Elige un remedio 🌿     ║")
    print("╚══════════════════════════════╝\n")
    print("1. Agua de vidrio")
    print("2. Concentrado de chile")
    print("3. Fungicida de ajo y vinagre")
    print("4. Agua con vinagre")
    opcion = int(input("¡Elige! "))

    if opcion == 1:
        print("\n╔══════════════════════════════════════╗")
        print("║      🌱 RECETA: AGUA DE VIDRIO       ║")
        print("╚══════════════════════════════════════╝")
        print("Ingredientes:")
        print(" - 100 g de cal")
        print(" - 400 g de ceniza")
        print(" - 10 L de agua")
        print("Mezclar los ingredientes y dejar reposar.")
    elif opcion == 2:
        print("\n╔══════════════════════════════════════╗")
        print("║  🌶️  CONCENTRADO DE CHILE            ║")
        print("╚══════════════════════════════════════╝")
        print("Ingredientes:")
        print(" - 2 cucharadas de chile en polvo")
        print(" - 1 L de agua")
        print(" - 1 cucharada de jabón líquido")
        print("1. Licuar los ingredientes")
        print("2. Dejar reposar 1 día")
        print("3. Colar")
    elif opcion == 3:
        print("\n╔══════════════════════════════════════╗")
        print("║  🧄 FUNGICIDA DE AJO Y VINAGRE       ║")
        print("╚══════════════════════════════════════╝")
        print("Ingredientes:")
        print(" - 5 dientes de ajo")
        print(" - 2 cucharadas de vinagre")
        print(" - 1 litro de agua")
        print("Licuar y colar.")
    elif opcion == 4:
        print("\n╔══════════════════════════════════════╗")
        print("║  🍶 AGUA CON VINAGRE                 ║")
        print("╚══════════════════════════════════════╝")
        print(" - 1 litro de agua")
        print(" - 1 cucharada de vinagre")
        print(" - Mezclar bien.")
    else:
        print("Opción inválida.")


# ─────────────────────────────────────────────
#  MÓDULO DE FICHA DE PLANTA  ← NUEVO
# ─────────────────────────────────────────────
def ficha_planta(usuario):
    print("\n╔══════════════════════════════╗")
    print("║     🌱 Ficha de tu planta    ║")
    print("╚══════════════════════════════╝\n")

    # Mostrar lista de plantas disponibles agrupadas por tipo
    print("Plantas disponibles en el huerto:\n")
    for tipo, info in INFO_TIPO.items():
        nombres = ", ".join(p.capitalize() for p in info["especies"])
        print(f"  {tipo}: {nombres}")

    nombre = input("\n¿Qué planta tienes? ").strip().lower()

    if nombre not in PLANTA_TIPO:
        print(f"\n❌ '{nombre.capitalize()}' no está en la base de datos del huerto.")
        return

    tipo = PLANTA_TIPO[nombre]
    info = INFO_TIPO[tipo]

    # ── Mostrar ficha de recomendaciones ──
    print(f"\n╔══════════════════════════════════════╗")
    print(f"║  📋 {nombre.capitalize():<33}║")
    print(f"╚══════════════════════════════════════╝")
    print(f"  Tipo             : {tipo}")
    print(f"  ☀️  Sol           : {info['sol']}")
    print(f"  💧 Riego         : {info['agua_veces']} / semana, {info['agua_min']} por vez")
    print(f"  📐 Densidad      : {info['plantas_m2']} plantas/m²")
    print(f"  ↔️  Separación    : {info['separacion']} entre plantas")

    # ── Dimensiones del huerto ──
    print()
    try:
        ancho = float(input("¿Ancho de tu espacio para esta planta (m)? "))
        largo = float(input("¿Largo de tu espacio para esta planta (m)? "))
    except ValueError:
        print("Medidas inválidas.")
        return

    area = round(ancho * largo, 2)

    rango  = info["plantas_m2"].replace("–", "-").split("-")
    sep_cm = int("".join(filter(str.isdigit, info["separacion"])))
    min_p  = int(float(rango[0].strip()) * area)
    max_p  = int(float(rango[1].strip()) * area)
    surcos = max(1, int((ancho * 100) / sep_cm))

    print(f"\n  📏 Área          : {area} m²")
    print(f"  🌿 Plantas       : entre {min_p} y {max_p} {nombre}s")
    print(f"  🪴 Surcos        : aprox. {surcos} surcos")
    print(f"     Por surco     : entre {min_p // surcos} y {max_p // surcos} plantas")

    # ── Malestar opcional ──
    print()
    tiene_malestar = input("¿Quieres registrar algún malestar hoy? (Si/No) ").strip().capitalize()
    malestar_texto = ""
    if tiene_malestar == "Si":
        print("\nVamos al diagnóstico...\n")
        malestar_texto = malestar()

    # ── Guardar en historial ──
    guardar_historial(usuario, nombre, tipo, ancho, largo, area, malestar_texto)


# ─────────────────────────────────────────────
#  MÓDULO DE GUARDADO EN HISTORIAL  ← NUEVO
# ─────────────────────────────────────────────
def guardar_historial(usuario, planta, tipo, ancho, largo, area, malestar_texto):
    nueva = {
        "usuario":  usuario,
        "planta":   planta.capitalize(),
        "tipo":     tipo,
        "fecha":    date.today().strftime("%d/%m/%Y"),
        "ancho":    ancho,
        "largo":    largo,
        "area":     area,
        "malestar": malestar_texto if malestar_texto else "Ninguno"
    }
    try:
        df_hist = pd.read_excel(ARCHIVO, sheet_name="Historial")
    except Exception:
        df_hist = pd.DataFrame(columns=nueva.keys())

    df_hist = pd.concat([df_hist, pd.DataFrame([nueva])], ignore_index=True)

    wb = load_workbook(ARCHIVO)
    if "Historial" in wb.sheetnames:
        del wb["Historial"]
    ws = wb.create_sheet("Historial")

    # Encabezados
    ws.append(list(df_hist.columns))
    for _, row in df_hist.iterrows():
        ws.append(list(row))

    wb.save(ARCHIVO)
    print("\n✅ Registro guardado en tu historial.\n")


# ─────────────────────────────────────────────
#  MÓDULO DE VER HISTORIAL  ← NUEVO
# ─────────────────────────────────────────────
def ver_historial(usuario):
    print("\n╔══════════════════════════════╗")
    print("║      📋 Tu historial         ║")
    print("╚══════════════════════════════╝\n")
    try:
        df_hist = pd.read_excel(ARCHIVO, sheet_name="Historial")
        mis = df_hist[df_hist["usuario"] == usuario]
        if mis.empty:
            print("  Aún no tienes registros guardados.\n")
        else:
            for _, r in mis.iterrows():
                print(f"  📅 {r['fecha']}  |  🌿 {r['planta']} ({r['tipo']})")
                print(f"     Espacio : {r['ancho']} m × {r['largo']} m  →  {r['area']} m²")
                print(f"     Malestar: {r['malestar']}\n")
    except Exception:
        print("  No hay historial guardado todavía.\n")


# ─────────────────────────────────────────────
#  MÓDULO DE FICHA DEL HUERTO
# ─────────────────────────────────────────────
def mostrar_ficha(usuario, area, ancho, largo):
    hoy = date.today().strftime("%d/%m/%Y")
    print("\n╔══════════════════════════════════╗")
    print("║         Ficha del huerto         ║")
    print("╚══════════════════════════════════╝")
    print(f"  Usuario      : {usuario}")
    print(f"  Fecha        : {hoy}")
    print(f"  Dimensiones  : {ancho} m × {largo} m")
    print(f"  Área total   : {area} m²\n")


# ─────────────────────────────────────────────
#  MENÚ PRINCIPAL
# ─────────────────────────────────────────────
def menu_principal(usuario):
    while True:
        print("╔══════════════════════════════════╗")
        print("║   ¿Ahora qué quieres hacer? 🌿   ║")
        print("╚══════════════════════════════════╝\n")
        print("1. Diagnosticar mi planta")
        print("2. Ficha de una planta del huerto")
        print("3. Ver mi historial de plantas")
        print("4. Preparar un remedio")
        print("5. Planear el acomodo de mi huerto")
        print("0. Salir")

        try:
            opcion = int(input("\n¡Elige! "))
        except ValueError:
            print("Por favor ingresa un número.\n")
            continue

        if opcion == 1:
            malestar()
        elif opcion == 2:
            ficha_planta(usuario)
        elif opcion == 3:
            ver_historial(usuario)
        elif opcion == 4:
            recetario()
        elif opcion == 5:
            calcular_plantas()
        elif opcion == 0:
            print("¡Hasta luego! 🌱")
            break
        else:
            print("Selección inválida.\n")


# ─────────────────────────────────────────────
#  PUNTO DE ENTRADA
# ─────────────────────────────────────────────
usuario = login()
menu_principal(usuario)
